from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
import pytest
import os
from openpyxl import Workbook, load_workbook

class DynamicTable:
    
    def __init__(self, driver):
        self.driver = driver
        
    def dynamic_table(self):
        driver = self.driver
        
        wait = WebDriverWait(driver, 10)
        
        dynamic_table_url = wait.until(EC.element_to_be_clickable((By.XPATH, "//div/h3/a[.='Dynamic Table']")))
        dynamic_table_url.click()
        time.sleep(2)
        
        wait.until(EC.visibility_of_element_located((By.XPATH, "//div[@role='rowgroup']/div[@role='row']")))
        list_of_names = driver.find_elements(By.XPATH, "//div[@role='rowgroup']/div[@role='row']")
        for i,item in enumerate(list_of_names):
            row = item.find_element(By.XPATH, ".//span[1]").text
            print(f"Item [{i}] : {row}")
        
            
        for i,item in enumerate(list_of_names):
            row = item.find_element(By.XPATH, ".//span[2]").text
            print(f"Item [{i}] : {row}")
        
        for i,item in enumerate(list_of_names):
            row = item.find_element(By.XPATH, ".//span[3]").text
            print(f"Item [{i}] : {row}")
         
           
        for i,item in enumerate(list_of_names):
            row = item.find_element(By.XPATH, ".//span[4]").text
            print(f"Item [{i}] : {row}")
        
        for i,item in enumerate(list_of_names):
            row = item.find_element(By.XPATH, ".//span[5]").text
            print(f"Item [{i}] : {row}")
            
        time.sleep(2)
        
    def extract_to_excel(self):
        driver = self.driver
        wait = WebDriverWait(driver, 10)
        
        file_name = "Extracted_browserinfo.xlsx"
        
        #if file exists, load it. Else, create a new workbook
        if os.path.exists(file_name):
            workbook = load_workbook(file_name)
        else:
            workbook = Workbook()
            default_sheet = workbook.active
            workbook.remove(default_sheet)
            
        sheet_title = "Browsers_Info"
        
        #avoid duplicate sheet names
        #if sheet already exists (e.g running multiple times)
        #add a number suffix
        base_title = sheet_title
        count = 1
        while sheet_title in workbook.sheetnames:
            count+=1
            sheet_title = f"{base_title} ({count})"
            
        #add new sheet
        sheet = workbook.create_sheet(title=sheet_title)
        
        #to append the column headers to the excel file, the column headers have to be a list of the column names
        wait.until(EC.visibility_of_element_located((By.XPATH, "//div[@role='rowgroup']/div[@role='row']/span[@role='columnheader']")))
        column_headers = driver.find_elements(By.XPATH, "//div[@role='rowgroup']/div[@role='row']/span[@role='columnheader']")
        col_headers_list = []
        
        for element in column_headers:
            col_headers_list.append(element.text)
            
        sheet.append(col_headers_list)
        
        rows = driver.find_elements(By.XPATH, "//div[@role='rowgroup']/div[@role='row']")
        for row in rows[1:]:
            cols = row.find_elements(By.XPATH, ".//span[@role='cell']")
            data = [col.text for col in cols]
            print(row.text)
            sheet.append(data)

        
        workbook.save("Extracted_browserinfo.xlsx")
        print(f"Table data saved to '{file_name}' in sheet '{sheet_title}'.")
        print("Current sheets in workbook: ", workbook.sheetnames)
            
        time.sleep(2)
        
    def take_screenshots(self):
        driver = self.driver
        wait = WebDriverWait(driver, 10)
        
        wait.until(EC.visibility_of_element_located((By.XPATH, "//div[@role='rowgroup']/div[@role='row']")))
        
        #make a folder to store screenshots (only if it doesn't exist)
        #step1: naming a folder where you'll save your screenshots
        screenshots_dir = "dynamic_table_screenshots"
        os.makedirs(screenshots_dir, exist_ok=True)
        
        #generate filename with timestamp
        timestamp = time.strftime("%Y-%m-%d_%H-%M-%S")
        image = driver.find_element(By.XPATH, "//div[@role='table']")
        filename = f"{screenshots_dir}/dynamic_table_{timestamp}.png"
        image.screenshot(filename)        
        time.sleep(2)
        
        print(f"Screenshot saved as {filename}")